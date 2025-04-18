import gradio as gr
import pandas as pd
from llama_index.core.agent.workflow import AgentWorkflow, ToolCallResult, AgentOutput, ToolCall
from llama_index.core.memory import ChatMemoryBuffer
from llama_index.core.storage.chat_store import SimpleChatStore
from openpyxl import load_workbook

from agents.markdown_table_agent import MarkdownTableAgent
from export_tools import export_to_markdown
from openai_like_llm import OpenAILikeLLM, OPENAI_MODEL_NAME, OPENAI_API_BASE, OPENAI_API_KEY
from tools.table_tool import clear_sheets_db, set_sheets_db, get_excel_info_tool, \
    get_all_table_names, get_sheets_db, is_regular_table, test_run_sql_queries

# logging.basicConfig(level="DEBUG")


llm = OpenAILikeLLM(model=OPENAI_MODEL_NAME, api_base=OPENAI_API_BASE, api_key=OPENAI_API_KEY)

# 是否上传了文档
is_uploaded = False

chat_store = SimpleChatStore()
chat_memory = ChatMemoryBuffer.from_defaults(
    token_limit=8000,
    chat_store=chat_store,
    chat_store_key="user",
)


async def analyze_question(question):
    global is_uploaded
    if not is_uploaded:
        gr.Warning("请先上传Excel文件")
        return "请先上传Excel文件"
    markdown_agent = MarkdownTableAgent(llm)
    agent_workflow = AgentWorkflow(
        agents=[markdown_agent.get_agent()],
        root_agent=markdown_agent.get_agent_name(),
    )

    handler = agent_workflow.run(
        user_msg=question,
        memory=chat_memory
    )
    current_agent = None
    final_output = ""
    async for event in handler.stream_events():
        if (
                hasattr(event, "current_agent_name")
                and event.current_agent_name != current_agent
        ):
            current_agent = event.current_agent_name
            print(f"\n{'=' * 50}")
            print(f"🤖 Agent: {current_agent}")
            print(f"{'=' * 50}\n")
        elif isinstance(event, AgentOutput):
            if event.response.content:
                print("📤 Output:", event.response.content)
                final_output += event.response.content
            if event.tool_calls:
                print(
                    "🛠️  Planning to use tools:",
                    [call.tool_name for call in event.tool_calls],
                )
        elif isinstance(event, ToolCallResult):
            print(f"🔧 Tool Result ({event.tool_name}):")
            print(f"  Arguments: {event.tool_kwargs}")
            print(f"  Output: {event.tool_output}")
        elif isinstance(event, ToolCall):
            print(f"🔨 Calling Tool: {event.tool_name}")
            print(f"  With arguments: {event.tool_kwargs}")
    return final_output


def load_excel(file):
    global is_uploaded
    # 清除之前加载的数据
    # 清除 sheets_db
    # clear_sheets_db()
    # 清除动态创建的全局变量
    table_names = list(get_all_table_names(get_sheets_db()))
    print(table_names)
    clear_sheets_db()

    # 使用 openpyxl 加载 Excel 文件
    wb = load_workbook(file.name)
    sheets_db = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        merged_regions = []

        # 收集所有合并区域信息（左上角坐标和区域范围）
        for merged_range in sheet.merged_cells.ranges:
            min_row, max_row = merged_range.min_row, merged_range.max_row
            min_col, max_col = merged_range.min_col, merged_range.max_col
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            merged_regions.append({
                'min_row': min_row,
                'max_row': max_row,
                'min_col': min_col,
                'max_col': max_col,
                'value': top_left_value
            })

        # 检查是否是空表（至少需要两行数据：标题行+数据行）
        if sheet.max_row <= 1:
            print(f"表 {sheet_name} 数据不足，拒绝加载。")
            continue

        # 构建处理后的数据矩阵
        data_matrix = []
        for row_idx in range(1, sheet.max_row + 1):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell_value = None
                # 检查当前单元格是否属于某个合并区域
                for region in merged_regions:
                    if (region['min_row'] <= row_idx <= region['max_row'] and
                            region['min_col'] <= col_idx <= region['max_col']):
                        cell_value = region['value']
                        break  # 找到所属区域后停止搜索
                if cell_value is None:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value
                row_data.append(cell_value)
            data_matrix.append(row_data)

        # 提取标题和数据
        headers = data_matrix[0]
        data_rows = data_matrix[1:]

        # 转换为DataFrame
        df = pd.DataFrame(data_rows, columns=headers)

        markdown_text = df.to_markdown()
        print(markdown_text)

        if not is_regular_table(df):
            print(f"表 {sheet_name} 不是正规表格，拒绝加载。")
            gr.Warning(f"表 {sheet_name} 不是正规表格，拒绝加载。")
            continue
        else:
            sheets_db[sheet_name] = df
            print(f"成功加载表 {sheet_name}，行数: {len(df)}")

        if not test_run_sql_queries(sheets_db):
            print(f"表 {sheet_name} 测试执行失败，拒绝加载。")
            gr.Warning(f"表 {sheet_name} 测试执行失败，拒绝加载。")
            return "表 {sheet_name} 测试执行失败，拒绝加载。"

    if not sheets_db:
        gr.Warning("没有找到正规表格，拒绝加载。")
        print("没有找到正规表格，拒绝加载。")
        return "没有找到正规表格，拒绝加载。"

    set_sheets_db(sheets_db)
    print(f"成功加载 {len(sheets_db)} 个工作表: {', '.join(get_all_table_names(sheets_db))}")

    info_str = get_excel_info_tool()
    # 打印 DataFrame 的信息和前几行数据
    print(info_str)
    is_uploaded = True

    return "Excel 文件已成功加载。"


with gr.Blocks() as excel_view:
    gr.Markdown("### Excel 表格分析系统")
    with gr.Row():
        with gr.Column():
            file_input = gr.File(label="选择 Excel 文件")
            load_output = gr.Textbox(label="文件加载结果")
            question_input = gr.Textbox(label="请输入问题", placeholder="输入你的问题")
        with gr.Column():
            answer_output = gr.Markdown(label="分析结果")
            # Replace Spinner with a hidden textbox to simulate loading state
            loading_indicator = gr.Textbox(visible=False, value="Loading...")
            # 添加 文件 导出按钮
            export_button = gr.Button("导出为 Markdown")
            export_button.click(
                fn=export_to_markdown,
                inputs=answer_output,
                outputs=gr.File(label="导出的 Markdown 文件")
            )

    file_input.upload(load_excel, inputs=file_input, outputs=load_output)
    # Modify the submit call to add loading state control
    question_input.submit(
        fn=analyze_question,
        inputs=question_input,
        outputs=answer_output,
        queue=True
    )

if __name__ == "__main__":
    excel_view.launch()
