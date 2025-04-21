import pandas as pd
import xlrd
from openpyxl import load_workbook


def load_excel_from_file(file_path, merge_cells=False):
    sheets_db = {}  # {sheet_name: DataFrame}
    # 判断文件扩展名
    if file_path.endswith('.xlsx') or file_path.endswith('.xlsm') or file_path.endswith('.xltx') or file_path.endswith(
            '.xltm'):
        # 使用 openpyxl 打开 .xlsx 等格式文件
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            print("total rows:", sheet.max_row)
            data = []
            header = None
            row_index = 0
            for row in sheet.iter_rows(values_only=True):
                if row_index == 0:
                    header = row
                    row_index += 1
                    continue
                data.append(row)
            df = pd.DataFrame(data, columns=header)
            sheets_db[sheet_name] = df
        return sheets_db
    elif file_path.endswith('.xls'):
        # 使用 xlrd 打开 .xls 格式文件
        workbook = xlrd.open_workbook(file_path)
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            print("total rows:", sheet.nrows)
            df = None
            # 读取表头
            header = sheet.row_values(0)
            data = []
            for row in range(1, sheet.nrows):  # 从第二行开始读取数据
                data.append(sheet.row_values(row))
            df = pd.DataFrame(data, columns=header)  # 设置表头
            sheets_db[sheet_name] = df
            return sheets_db
    else:
        print("不支持的文件格式，请使用 .xls 或 .xlsx 格式的文件。")
        # 抛出异常
        raise ValueError("不支持的文件格式，请使用.xls 或.xlsx 格式的文件。")